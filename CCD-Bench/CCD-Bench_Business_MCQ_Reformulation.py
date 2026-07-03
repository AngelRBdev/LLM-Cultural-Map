import pandas as pd, re, random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

random.seed(42)
GLOBES=['anglo','eastern_europe','latin-america','latin_europe','confucian_asia',
        'nordic_europe','sub-saharan_africa','southern-asia','germanic_europe','middle_east']
SOC_ORDER=GLOBES[:]

# Cultural orientation per globe on the individualism<->collectivism axis (from GLOBE In-Group/Institutional Collectivism).
INDIV = {'anglo','germanic_europe','nordic_europe'}
COLL  = {'confucian_asia','southern-asia','middle_east','latin-america','sub-saharan_africa','eastern_europe','latin_europe'}

import ast
def parse_list(c):
    try:
        v=ast.literal_eval(c); return v if isinstance(v,list) else []
    except: return []

soc=pd.read_excel('/mnt/user-data/uploads/Globe_Societies_Extended.xlsx')
gc={}
for i,g in enumerate(SOC_ORDER):
    ctr=parse_list(soc.iloc[i]['Societies'])+parse_list(soc.iloc[i]['Extended_Societies'])
    ctr=[re.sub(r'\s*\([^)]*\)','',c).strip() for c in ctr]
    seen=set(); pool=[]
    for c in ctr:
        if c not in seen: seen.add(c); pool.append(c)
    gc[g]=pool

df=pd.read_excel('/mnt/user-data/uploads/CCD-Bench_Business_V2.xlsx')

# Detect polarity of Option 1: is it the individualist pole?
INDIV_WORDS={'personal','individual','own','self','myself','independently','independent','my ','i would','autonomy','alone','solo','personally'}
COLL_WORDS={'team','group','collective','together','shared','harmony','everyone','family','community','our ','company','colleagues','collaborat'}
def pole(text):
    t=' '+str(text).lower()+' '
    iv=sum(t.count(w) for w in INDIV_WORDS)
    cv=sum(t.count(w) for w in COLL_WORDS)
    return 'indiv' if iv>=cv else 'coll'

# country-anchored question reword: replace 2nd person with 3rd person about an employee in COUNTRY
def reword_q(q, country):
    q=str(q).strip()
    # strip leading 2nd-person framings to avoid "How would an employee... How would you..."
    q=re.sub(r'^(how would you|how do you|what would be your|what is your|when you|in your|do you|would you|which|what)\b.*?\?\s*','',q,flags=re.I) if q.count('?')>1 else q
    q2=re.sub(r'\byour\b','their',q,flags=re.I); q2=re.sub(r'\byou\b','they',q2,flags=re.I)
    return f"How would an employee in {country} handle this? {q2}"
def reword_opt(o):
    o=str(o).strip()
    o=re.sub(r'\byour\b','their',o,flags=re.I)
    o=re.sub(r'\byou\b','they',o,flags=re.I)
    return o[0].upper()+o[1:] if o else o

DISTRACTORS=[
 "Seek a negotiated compromise that partially satisfies both their own position and the opposing one.",
 "Defer the decision to a senior authority or supervisor and follow whatever ruling is given.",
 "Postpone acting until they have gathered more input from trusted advisors before committing.",
 "Adopt whichever option carries the least personal risk regardless of the underlying principle.",
]

rows=[]; qid=0
for _,r in df.iterrows():
    qid+=1
    o1=reword_opt(r['Option 1']); o2=reword_opt(r['Option 2'])
    p1=pole(r['Option 1'])  # polarity of option 1
    for g in GLOBES:
        orient='indiv' if g in INDIV else 'coll'
        # correct = option whose polarity matches the globe's orientation
        if p1==orient:
            correct_text=o1
        else:
            correct_text=o2
        country=random.choice(gc[g]) if gc[g] else g
        d=random.sample(DISTRACTORS,2)
        opts=[o1,o2,d[0],d[1]]
        random.shuffle(opts)
        ci=opts.index(correct_text)
        rows.append({
            'q_id':qid,'domain':str(r['Domain']),'globe':g,'country':country,
            'question':reword_q(r['Question'],country),
            'option_A':opts[0],'option_B':opts[1],'option_C':opts[2],'option_D':opts[3],
            'correct_option':['A','B','C','D'][ci],
            'orientation':orient,'original_question':str(r['Question'])
        })

out=pd.DataFrame(rows)
print('Items:',len(out),'| source Qs:',out['q_id'].nunique())
print('Correct-option balance:',out['correct_option'].value_counts().to_dict())
print('Orientation split:',out['orientation'].value_counts().to_dict())

wb=Workbook(); ws=wb.active; ws.title='Benchmark_MCQ'
H=['q_id','domain','globe','country','question','option_A','option_B','option_C','option_D','correct_option','original_question']
ws.append(H)
for _,r in out.iterrows(): ws.append([r[h] for h in H])
hf=PatternFill('solid',start_color='1F4E78'); hfont=Font(bold=True,color='FFFFFF',name='Arial')
thin=Side(style='thin',color='D9D9D9'); bd=Border(left=thin,right=thin,top=thin,bottom=thin)
for c in range(1,len(H)+1):
    cell=ws.cell(row=1,column=c); cell.fill=hf; cell.font=hfont
    cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
for col,w in {'A':7,'B':12,'C':16,'D':16,'E':55,'F':42,'G':42,'H':42,'I':42,'J':10,'K':50}.items():
    ws.column_dimensions[col].width=w
ws.freeze_panes='A2'
for row in ws.iter_rows(min_row=2,max_col=len(H)):
    for cell in row:
        cell.font=Font(name='Arial',size=10); cell.alignment=Alignment(vertical='top',wrap_text=True); cell.border=bd
for r in range(2,ws.max_row+1):
    ws.cell(row=r,column=10).fill=PatternFill('solid',start_color='E2EFDA')
    ws.cell(row=r,column=10).alignment=Alignment(horizontal='center',vertical='center')

ws3=wb.create_sheet('README_methodology')
for n in [
 ['CCD-Bench Business -> 4-option MCQ, country-anchored, per GLOBE'],[''],
 ['Format follows Ahmed example: Option 1/2 = original two options (reworded to'],
 ['3rd person, country-anchored). Option 3/4 = generated plausible distractors'],
 ['(compromise / defer-to-authority). One correct answer per item.'],[''],
 ['CORRECT ANSWER derivation (key method):'],
 ['Each item is scored on the individualism<->collectivism axis. The correct'],
 ['option is the one whose polarity matches the GLOBE orientation of the anchor'],
 ['cluster, from In-Group/Institutional Collectivism scores:'],
 ['  Individualist -> Option 1-type: anglo, germanic_europe, nordic_europe'],
 ['  Collectivist  -> Option 2-type: confucian_asia, southern-asia, middle_east,'],
 ['     latin-america, sub-saharan_africa, eastern_europe, latin_europe'],
 ['Polarity of Option 1 is detected per row by lexicon (not assumed constant),'],
 ['so the label is not inverted when Option 1 happens to be the collectivist pole.'],[''],
 ['CAVEATS to disclose:'],
 ['1. Method assumes every question reduces to the individ/collect axis. Questions'],
 ['   that hinge on other GLOBE dimensions (Power Distance, Uncertainty Avoidance)'],
 ['   may be mislabelled. Manual audit recommended on a sample.'],
 ['2. Correct answers are NOT taken from the globe-answer cells in V2, which were'],
 ['   found to be row-misaligned with their questions. They are derived from'],
 ['   documented cluster profiles instead.'],
 ['3. Distractors are generic dilemma-domain options, not real cluster answers.'],
]: ws3.append(n)
ws3.column_dimensions['A'].width=82; ws3.cell(row=1,column=1).font=Font(bold=True,size=13,name='Arial')
wb.save('/home/claude/CCD-Bench_MCQ_v2.xlsx'); print('saved')
